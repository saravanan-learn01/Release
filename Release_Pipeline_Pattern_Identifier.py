"""
Developer: Kolappan Mayilvanan
Script: Release_Pipeline_Pattern_Identifier.py
Version: 3.0 - Excel Output with Dual Sheets
Description: This script identifies release pipeline JSON files with similar task patterns.
             Creates Excel file with:
             Sheet 1: Matched patterns (Pipeline and its matching pipelines)
             Sheet 2: Unique pipelines (Pipelines with no matches)
"""

import os
import json
import sys
from datetime import datetime
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def print_header(message):
    """Print formatted header"""
    print("\n" + "=" * 70)
    print(f" {message}")
    print("=" * 70)

def print_section(message):
    """Print section header"""
    print("\n" + "-" * 50)
    print(f" {message}")
    print("-" * 50)

def load_json(file_path):
    """Load JSON file and return data or error"""
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            return json.load(file), None
    except json.JSONDecodeError as e:
        return None, f"JSON Decode Error: {str(e)}"
    except Exception as e:
        return None, str(e)

def extract_task_ids_from_release_pipeline(data):
    """
    Extract task IDs from release pipeline JSON structure
    Structure: environments → deployPhases → workflowTasks → taskId
    """
    task_ids = []
    
    if not isinstance(data, dict):
        return None
    
    environments = data.get('environments', [])
    if not environments:
        return None
    
    has_any_tasks = False
    
    for environment in environments:
        deploy_phases = environment.get('deployPhases', [])
        
        for phase in deploy_phases:
            workflow_tasks = phase.get('workflowTasks', [])
            
            if workflow_tasks:
                has_any_tasks = True
                
            for task_item in workflow_tasks:
                # Try different ways to get task ID
                task_id = task_item.get('taskId')
                
                if not task_id and 'task' in task_item:
                    task_obj = task_item.get('task', {})
                    task_id = task_obj.get('id')
                
                if task_id:
                    task_ids.append(task_id)
    
    if not has_any_tasks:
        return None
    
    return task_ids

def extract_task_details_from_release_pipeline(data):
    """
    Extract detailed task information including names and versions
    """
    task_details = []
    
    if not isinstance(data, dict):
        return None
    
    environments = data.get('environments', [])
    
    for environment in environments:
        env_name = environment.get('name', 'Unknown')
        deploy_phases = environment.get('deployPhases', [])
        
        for phase in deploy_phases:
            workflow_tasks = phase.get('workflowTasks', [])
            
            for task_item in workflow_tasks:
                task_id = task_item.get('taskId')
                task_name = task_item.get('name', '')
                task_version = task_item.get('version', '')
                
                if task_id:
                    task_details.append({
                        'task_id': task_id,
                        'task_name': task_name,
                        'task_version': task_version,
                        'environment': env_name
                    })
    
    return task_details if task_details else None

def find_json_files(directory):
    """Recursively find all JSON files in directory"""
    json_files = []
    for root, _, files in os.walk(directory):
        for file in files:
            if file.endswith('.json'):
                json_files.append(os.path.join(root, file))
    return json_files

def parse_command_line_arguments():
    """Parse and validate command line arguments"""
    if len(sys.argv) < 2:
        print("\nError: Input directory path is required!")
        print("\nUsage: python script.py <directory_path> [options]")
        print("\nRequired Arguments:")
        print("  <directory_path>        : Path to directory containing release pipeline JSON files")
        print("\nOptional Arguments:")
        print("  --output_excel <filename> : Custom output Excel filename (default: Release_Pipeline_Patterns.xlsx)")
        print("  --error_log <filename>    : Custom error log filename (default: error_log.txt)")
        print("  --no_match_log <filename> : Custom no-match log filename (default: no_tasks_log.txt)")
        print("  --debug                    : Show detailed task information")
        print("  --help                     : Show this help message")
        return None, None, None, None, False
    
    directory = sys.argv[1]
    output_file = "Release_Pipeline_Patterns.xlsx"
    error_log_file = "error_log.txt"
    no_match_log_file = "no_tasks_log.txt"
    debug_mode = False
    
    i = 2
    while i < len(sys.argv):
        if sys.argv[i] == '--output_excel' and i + 1 < len(sys.argv):
            output_file = sys.argv[i + 1]
            if not output_file.endswith('.xlsx'):
                output_file += '.xlsx'
            i += 2
        elif sys.argv[i] == '--error_log' and i + 1 < len(sys.argv):
            error_log_file = sys.argv[i + 1]
            i += 2
        elif sys.argv[i] == '--no_match_log' and i + 1 < len(sys.argv):
            no_match_log_file = sys.argv[i + 1]
            i += 2
        elif sys.argv[i] == '--debug':
            debug_mode = True
            i += 1
        elif sys.argv[i] == '--help':
            print_help()
            sys.exit(0)
        else:
            print(f"\nWarning: Unknown argument '{sys.argv[i]}' - ignoring")
            i += 1
    
    return directory, output_file, error_log_file, no_match_log_file, debug_mode

def print_help():
    """Print help information"""
    print("\nHelp: Release Pipeline Pattern Identifier - Excel Output")
    print("=" * 70)
    print("This script identifies release pipeline JSON files with similar task patterns.")
    print("Creates Excel file with two sheets:")
    print("  - Matched Patterns: Shows pipelines with identical task sets")
    print("  - Unique Pipelines: Shows pipelines with unique task sets")
    print("\nUsage: python script.py <directory_path> [options]")
    print("\nRequired Arguments:")
    print("  <directory_path>        : Path to directory containing release pipeline JSON files")
    print("\nOptional Arguments:")
    print("  --output_excel <filename> : Custom output Excel filename")
    print("  --error_log <filename>    : Custom error log filename")
    print("  --no_match_log <filename> : Custom no-match log filename")
    print("  --debug                    : Show detailed task information")
    print("  --help                     : Show this help message")
    print("\nExample:")
    print("  python script.py C:\\Output\\MyCollection --output_excel my_patterns.xlsx")

def analyze_files(file_task_ids, file_task_details, directory):
    """
    Analyze files and find patterns
    Returns matched_patterns and unique_files
    """
    # Group by exact task sequence
    pattern_dict = {}
    for file_path, task_ids in file_task_ids.items():
        task_tuple = tuple(task_ids)  # Keep order and duplicates
        if task_tuple not in pattern_dict:
            pattern_dict[task_tuple] = []
        pattern_dict[task_tuple].append(file_path)
    
    # Separate matched and unique
    matched_patterns = {}  # task_tuple -> list of files (only if >1 file)
    unique_files = []      # list of files with unique patterns
    
    for task_tuple, files in pattern_dict.items():
        if len(files) > 1:
            matched_patterns[task_tuple] = files
        else:
            unique_files.append(files[0])
    
    return matched_patterns, unique_files

def apply_excel_styles(ws):
    """Apply formatting to Excel worksheet"""
    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Apply header styles
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 100)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_matched_patterns_sheet(wb, matched_patterns, file_task_details, directory):
    """Create worksheet for matched patterns"""
    ws = wb.create_sheet("Matched Patterns")
    
    # Headers
    headers = ['Pattern ID', 'Pipeline Name', 'Pipeline Path', 'Task Count', 'Task IDs', 'Task Names']
    ws.append(headers)
    
    row_num = 2
    pattern_id = 1
    
    for task_tuple, files in matched_patterns.items():
        # For each file in the pattern group
        for file_path in files:
            rel_path = os.path.relpath(file_path, directory)
            task_details = file_task_details.get(file_path, [])
            task_count = len(task_details)
            task_ids = ', '.join([t['task_id'] for t in task_details][:5])
            if task_details and len(task_details) > 5:
                task_ids += f" ... ({len(task_details)} total)"
            
            task_names = ', '.join(list(set([t['task_name'] for t in task_details if t['task_name']]))[:3])
            if task_details:
                unique_names = set([t['task_name'] for t in task_details if t['task_name']])
                if len(unique_names) > 3:
                    task_names += f" ... ({len(unique_names)} total)"
            
            ws.cell(row=row_num, column=1, value=f"Pattern {pattern_id}")
            ws.cell(row=row_num, column=2, value=os.path.basename(file_path))
            ws.cell(row=row_num, column=3, value=rel_path)
            ws.cell(row=row_num, column=4, value=task_count)
            ws.cell(row=row_num, column=5, value=task_ids)
            ws.cell(row=row_num, column=6, value=task_names)
            row_num += 1
        
        pattern_id += 1
    
    # Add summary at the bottom
    if len(matched_patterns) > 0:
        ws.append([])
        ws.append(['SUMMARY'])
        ws.append(['Total Patterns Found', len(matched_patterns)])
        ws.append(['Total Files in Patterns', sum(len(files) for files in matched_patterns.values())])
    
    apply_excel_styles(ws)

def create_unique_pipelines_sheet(wb, unique_files, file_task_details, directory):
    """Create worksheet for unique pipelines"""
    ws = wb.create_sheet("Unique Pipelines")
    
    # Headers
    headers = ['Pipeline Name', 'Pipeline Path', 'Task Count', 'Task IDs', 'Task Names']
    ws.append(headers)
    
    row_num = 2
    
    for file_path in sorted(unique_files):
        rel_path = os.path.relpath(file_path, directory)
        task_details = file_task_details.get(file_path, [])
        task_count = len(task_details)
        task_ids = ', '.join([t['task_id'] for t in task_details][:5])
        if task_details and len(task_details) > 5:
            task_ids += f" ... ({len(task_details)} total)"
        
        task_names = ', '.join(list(set([t['task_name'] for t in task_details if t['task_name']]))[:3])
        if task_details:
            unique_names = set([t['task_name'] for t in task_details if t['task_name']])
            if len(unique_names) > 3:
                task_names += f" ... ({len(unique_names)} total)"
        
        ws.cell(row=row_num, column=1, value=os.path.basename(file_path))
        ws.cell(row=row_num, column=2, value=rel_path)
        ws.cell(row=row_num, column=3, value=task_count)
        ws.cell(row=row_num, column=4, value=task_ids)
        ws.cell(row=row_num, column=5, value=task_names)
        row_num += 1
    
    # Add summary at the bottom
    if len(unique_files) > 0:
        ws.append([])
        ws.append(['SUMMARY'])
        ws.append(['Total Unique Pipelines', len(unique_files)])
    
    apply_excel_styles(ws)

def compare_json_files(directory, error_log_file, no_match_log_file, debug_mode=False):
    """
    Compare JSON files to find patterns
    """
    print_section("SCANNING DIRECTORY")
    print(f"Scanning directory: {directory}")
    
    json_files = find_json_files(directory)
    print(f"Found {len(json_files)} JSON file(s)")
    
    file_task_ids = {}      # Store file path -> list of task IDs
    file_task_details = {}   # Store file path -> detailed task info
    error_files = []         # Files with errors
    no_tasks_files = []      # Files with no tasks
    
    # First pass: Extract task information from all files
    print_section("EXTRACTING TASK INFORMATION")
    
    for json_file in json_files:
        rel_path = os.path.relpath(json_file, directory)
        print(f"  Processing: {rel_path}", end='', flush=True)
        
        data, error = load_json(json_file)
        if error:
            error_files.append(f"{rel_path}: {error}")
            print(f" - ERROR")
            continue
        
        task_ids = extract_task_ids_from_release_pipeline(data)
        task_details = extract_task_details_from_release_pipeline(data)
        
        if task_ids is None:
            no_tasks_files.append(rel_path)
            print(f" - No tasks found")
        else:
            file_task_ids[json_file] = task_ids
            file_task_details[json_file] = task_details
            unique_task_ids = set(task_ids)
            print(f" - {len(task_ids)} task(s) ({len(unique_task_ids)} unique)")
            
            if debug_mode:
                print(f"    Tasks: {', '.join(task_ids)}")
    
    print(f"\nSummary:")
    print(f"  Files with tasks: {len(file_task_ids)}")
    print(f"  Files with errors: {len(error_files)}")
    print(f"  Files with no tasks: {len(no_tasks_files)}")
    
    # Analyze patterns
    print_section("ANALYZING PATTERNS")
    matched_patterns, unique_files = analyze_files(file_task_ids, file_task_details, directory)
    
    print(f"\nResults:")
    print(f"  Found {len(matched_patterns)} pattern(s) with matching task sets")
    print(f"  Found {len(unique_files)} unique pipeline(s) with no matches")
    
    # Write error log
    if error_files:
        with open(error_log_file, 'w', encoding='utf-8') as error_log:
            error_log.write(f"Release Pipeline Pattern Identifier - Error Log\n")
            error_log.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            error_log.write(f"Directory: {directory}\n")
            error_log.write("=" * 70 + "\n\n")
            for error in error_files:
                error_log.write(f"{error}\n")
        print(f"\nError log written to: {error_log_file}")
    
    # Write no tasks log
    if no_tasks_files:
        with open(no_match_log_file, 'w', encoding='utf-8') as no_match_log:
            no_match_log.write(f"Release Pipeline Pattern Identifier - No Tasks Log\n")
            no_match_log.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            no_match_log.write(f"Directory: {directory}\n")
            no_match_log.write("=" * 70 + "\n\n")
            no_match_log.write("Files with no tasks found:\n")
            no_match_log.write("-" * 50 + "\n")
            for file_path in sorted(no_tasks_files):
                no_match_log.write(f"{file_path}\n")
            no_match_log.write(f"\nTotal files with no tasks: {len(no_tasks_files)}")
        print(f"\nNo tasks log written to: {no_match_log_file}")
    
    return matched_patterns, unique_files, file_task_details

def create_excel_output(matched_patterns, unique_files, file_task_details, directory, output_file):
    """Create Excel file with two sheets"""
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    
    # Create sheets
    if matched_patterns:
        create_matched_patterns_sheet(wb, matched_patterns, file_task_details, directory)
    
    if unique_files:
        create_unique_pipelines_sheet(wb, unique_files, file_task_details, directory)
    
    # If no sheets created (shouldn't happen), create at least one
    if not wb.sheetnames:
        ws = wb.create_sheet("No Data")
        ws.append(['No pipeline data found in the specified directory'])
    
    # Save the workbook
    wb.save(output_file)
    print(f"\n✓ Excel file created: {output_file}")

def main():
    """Main function"""
    
    print_header("RELEASE PIPELINE PATTERN IDENTIFIER - EXCEL OUTPUT")
    
    # Check for openpyxl
    try:
        from openpyxl import Workbook
    except ImportError:
        print("\nError: openpyxl is required for Excel output.")
        print("Install it using: pip install openpyxl")
        sys.exit(1)
    
    # Parse command line arguments
    directory, output_file, error_log_file, no_match_log_file, debug_mode = parse_command_line_arguments()
    
    if directory is None:
        sys.exit(1)
    
    # Validate directory
    if not os.path.exists(directory):
        print(f"Error: Directory '{directory}' does not exist.")
        sys.exit(1)
    if not os.path.isdir(directory):
        print(f"Error: '{directory}' is not a directory.")
        sys.exit(1)
    
    print(f"\nConfiguration:")
    print(f"  Input Directory : {directory}")
    print(f"  Output Excel    : {output_file}")
    print(f"  Error Log       : {error_log_file}")
    print(f"  No Tasks Log    : {no_match_log_file}")
    print(f"  Debug Mode      : {'Enabled' if debug_mode else 'Disabled'}")
    
    # Compare files
    matched_patterns, unique_files, file_task_details = compare_json_files(
        directory, error_log_file, no_match_log_file, debug_mode
    )
    
    # Create Excel output
    print_section("CREATING EXCEL REPORT")
    create_excel_output(matched_patterns, unique_files, file_task_details, directory, output_file)
    
    print_header("RESULTS SUMMARY")
    
    if matched_patterns:
        print(f"\n✓ Found {len(matched_patterns)} pattern(s) with matching pipelines")
        print(f"  - Check 'Matched Patterns' sheet in {output_file}")
        for i, (task_tuple, files) in enumerate(matched_patterns.items(), 1):
            print(f"    Pattern {i}: {len(files)} pipelines")
    else:
        print(f"\n✗ No matching patterns found")
    
    if unique_files:
        print(f"\n✓ Found {len(unique_files)} unique pipeline(s) with no matches")
        print(f"  - Check 'Unique Pipelines' sheet in {output_file}")
    else:
        print(f"\n✗ No unique pipelines found")
    
    print(f"\nExcel file location: {os.path.abspath(output_file)}")
    print("\n" + "=" * 70)

if __name__ == "__main__":
    main()