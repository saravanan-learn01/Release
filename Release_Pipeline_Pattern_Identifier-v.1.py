import os
import json
import sys
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Global variables for paths
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
LOGS_DIR = os.path.join(SCRIPT_DIR, "Logs")
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "Output")

def ensure_directories():
    """Create Logs and Output directories if they don't exist"""
    directories_created = []
    
    if not os.path.exists(LOGS_DIR):
        os.makedirs(LOGS_DIR)
        directories_created.append(f"Logs: {LOGS_DIR}")
    
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
        directories_created.append(f"Output: {OUTPUT_DIR}")
    
    if directories_created:
        print("✓ Created directories:")
        for dir_path in directories_created:
            print(f"  • {dir_path}")
    
    return LOGS_DIR, OUTPUT_DIR

def get_timestamp():
    """Get current timestamp for file naming"""
    return datetime.now().strftime('%Y%m%d_%H%M%S')

def print_header(message):
    """Print formatted header"""
    print("\n" + "=" * 70)
    print(f" {message}")
    print("=" * 70)

def print_section(message):
    """Print section header"""
    print("\n" + "-" * 50)
    print(f" {message}")
    print("-" * 50)

def load_json(file_path):
    """Load JSON file and return data or error"""
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            return json.load(file), None
    except json.JSONDecodeError as e:
        return None, f"JSON Decode Error: {str(e)}"
    except Exception as e:
        return None, str(e)

def extract_task_ids_from_release_pipeline(data):
    """
    Extract task IDs from release pipeline JSON structure
    """
    task_ids = []
    
    if not isinstance(data, dict):
        return None
    
    environments = data.get('environments', [])
    if not environments:
        return None
    
    has_any_tasks = False
    
    for environment in environments:
        deploy_phases = environment.get('deployPhases', [])
        
        for phase in deploy_phases:
            workflow_tasks = phase.get('workflowTasks', [])
            
            if workflow_tasks:
                has_any_tasks = True
                
            for task_item in workflow_tasks:
                task_id = task_item.get('taskId')
                if task_id:
                    task_ids.append(task_id)
    
    if not has_any_tasks:
        return None
    
    return task_ids

def extract_task_details_from_release_pipeline(data):
    """
    Extract detailed task information including names and versions
    """
    task_details = []
    
    if not isinstance(data, dict):
        return None
    
    environments = data.get('environments', [])
    
    for environment in environments:
        env_name = environment.get('name', 'Unknown')
        deploy_phases = environment.get('deployPhases', [])
        
        for phase in deploy_phases:
            workflow_tasks = phase.get('workflowTasks', [])
            
            for task_item in workflow_tasks:
                task_id = task_item.get('taskId')
                task_name = task_item.get('name', '')
                task_version = task_item.get('version', '')
                
                if task_id:
                    task_details.append({
                        'task_id': task_id,
                        'task_name': task_name,
                        'task_version': task_version,
                        'environment': env_name
                    })
    
    return task_details if task_details else None

def find_json_files(directory):
    """Recursively find all JSON files in directory"""
    json_files = []
    for root, _, files in os.walk(directory):
        for file in files:
            if file.endswith('.json'):
                json_files.append(os.path.join(root, file))
    return json_files

def parse_command_line_arguments():
    """
    Parse and validate command line arguments
    Format: python script.py --json_folder <folder_path>
    """
    if len(sys.argv) < 3 or '--json_folder' not in sys.argv:
        print("\n" + "=" * 70)
        print(" ERROR: Missing required argument --json_folder")
        print("=" * 70)
        print("\nUsage: python script.py --json_folder <folder_path>")
        print("\nRequired Arguments:")
        print("  --json_folder <folder_path> : Path to folder containing release pipeline JSON files")
        print("\nOptional Arguments:")
        print("  --help                      : Show this help message")
        print("\nExamples:")
        print("  python script.py --json_folder C:\\MyPipelines")
        print("  python script.py --json_folder ./json_files")
        print("\n" + "=" * 70)
        return None
    
    # Parse arguments
    directory = None
    i = 1
    while i < len(sys.argv):
        if sys.argv[i] == '--json_folder' and i + 1 < len(sys.argv):
            directory = sys.argv[i + 1]
            i += 2
        elif sys.argv[i] == '--help':
            print_help()
            sys.exit(0)
        else:
            print(f"\nWarning: Unknown argument '{sys.argv[i]}' - ignoring")
            i += 1
    
    return directory

def print_help():
    """Print help information"""
    print("\n" + "=" * 70)
    print(" RELEASE PIPELINE PATTERN IDENTIFIER - HELP")
    print("=" * 70)
    print("\nDescription:")
    print("  This script identifies release pipeline JSON files with similar task patterns.")
    print("\nFeatures:")
    print("  • Creates Excel file with two sheets:")
    print("    - Unique Pipelines: Pipelines with unique task patterns")
    print("    - Similar Pipelines: Pipelines that share the same task patterns")
    print("  • Output shows only Pipeline Name and Pipeline Path")
    print("\nUsage:")
    print("  python script.py --json_folder <folder_path>")
    print("\nRequired Arguments:")
    print("  --json_folder <folder_path> : Path to folder containing release pipeline JSON files")
    print("\nExamples:")
    print("  python script.py --json_folder C:\\MyPipelines")
    print("  python script.py --json_folder ./json_files")
    print("\nOutput:")
    print("  • Excel file: Output/Release_Pipeline_Patterns_YYYYMMDD_HHMMSS.xlsx")
    print("  • Logs folder: Logs/ with timestamped analysis files")
    print("\n" + "=" * 70)

def analyze_files(file_task_ids, file_task_details):
    """
    Analyze files and find patterns
    """
    # Group by exact task sequence
    pattern_dict = {}
    for file_path, task_ids in file_task_ids.items():
        task_tuple = tuple(task_ids)
        if task_tuple not in pattern_dict:
            pattern_dict[task_tuple] = []
        pattern_dict[task_tuple].append(file_path)
    
    # Separate matched and unique
    pattern_groups = {}
    unique_files = []
    
    for task_tuple, files in pattern_dict.items():
        if len(files) > 1:
            pattern_groups[task_tuple] = files
        else:
            unique_files.append(files[0])
    
    return pattern_groups, unique_files

def create_excel_output(pattern_groups, unique_files, file_task_details, directory, output_file):
    """Create Excel file with two sheets: Unique Pipelines and Similar Pipelines"""
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    
    # Create Unique Pipelines sheet
    create_unique_pipelines_sheet(wb, unique_files, file_task_details, directory)
    
    # Create Similar Pipelines sheet
    create_similar_pipelines_sheet(wb, pattern_groups, file_task_details, directory)
    
    # Save the workbook
    wb.save(output_file)
    print(f"\n✓ Excel file created: {output_file}")

def create_unique_pipelines_sheet(wb, unique_files, file_task_details, directory):
    """Create worksheet for unique pipelines with only Pipeline Name and Pipeline Path"""
    ws = wb.create_sheet("Unique Pipelines")
    
    # Headers - Only Pipeline Name and Pipeline Path
    headers = ['Pipeline Name', 'Pipeline Path']
    ws.append(headers)
    
    # Apply header styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    row_num = 2
    
    for file_path in sorted(unique_files):
        rel_path = os.path.relpath(file_path, directory)
        pipeline_name = os.path.basename(file_path)
        
        # Write to Excel - only the two columns
        ws.cell(row=row_num, column=1, value=pipeline_name)
        ws.cell(row=row_num, column=2, value=rel_path)
        
        # Add border to cells
        for col in range(1, 3):
            cell = ws.cell(row=row_num, column=col)
            cell.border = thin_border
        
        row_num += 1
    
    # Add summary at the bottom
    if unique_files:
        ws.append([])
        summary_row = row_num + 1
        ws.cell(row=summary_row, column=1, value="SUMMARY")
        ws.cell(row=summary_row + 1, column=1, value="Total Unique Pipelines")
        ws.cell(row=summary_row + 1, column=2, value=len(unique_files))
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 100)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_similar_pipelines_sheet(wb, pattern_groups, file_task_details, directory):
    """Create worksheet for similar pipelines with only Pattern ID, Pipeline Name, and Pipeline Path"""
    ws = wb.create_sheet("Similar Pipelines")
    
    # Headers - Only Pattern ID, Pipeline Name, and Pipeline Path
    headers = ['Pattern ID', 'Pipeline Name', 'Pipeline Path']
    ws.append(headers)
    
    # Apply header styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    row_num = 2
    pattern_id = 1
    
    for task_tuple, files in pattern_groups.items():
        # Add a pattern separator row (optional)
        pattern_start_row = row_num
        ws.cell(row=pattern_start_row, column=1, value=f"=== Pattern {pattern_id} ===")
        ws.cell(row=pattern_start_row, column=2, value=f"({len(files)} pipelines)")
        
        # Style the pattern header
        for col in range(1, 4):
            cell = ws.cell(row=pattern_start_row, column=col)
            cell.font = Font(bold=True, italic=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        row_num += 1
        
        # For each file in the pattern group
        for file_path in files:
            rel_path = os.path.relpath(file_path, directory)
            pipeline_name = os.path.basename(file_path)
            
            # Write to Excel - only the three columns
            ws.cell(row=row_num, column=1, value=f"Pattern {pattern_id}")
            ws.cell(row=row_num, column=2, value=pipeline_name)
            ws.cell(row=row_num, column=3, value=rel_path)
            
            # Add border to cells
            for col in range(1, 4):
                cell = ws.cell(row=row_num, column=col)
                cell.border = thin_border
            
            row_num += 1
        
        # Add an empty row between patterns
        row_num += 1
        pattern_id += 1
    
    # Add summary at the bottom
    if pattern_groups:
        ws.append([])
        summary_row = row_num + 1
        ws.cell(row=summary_row, column=1, value="SUMMARY")
        ws.cell(row=summary_row + 1, column=1, value="Total Patterns Found")
        ws.cell(row=summary_row + 1, column=2, value=len(pattern_groups))
        ws.cell(row=summary_row + 2, column=1, value="Total Pipelines in Patterns")
        total_pipelines = sum(len(files) for files in pattern_groups.values())
        ws.cell(row=summary_row + 2, column=2, value=total_pipelines)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 100)
        ws.column_dimensions[column_letter].width = adjusted_width

def compare_json_files(directory):
    """
    Compare JSON files to find patterns
    """
    print_section("SCANNING DIRECTORY")
    print(f"Scanning directory: {directory}")
    
    json_files = find_json_files(directory)
    print(f"Found {len(json_files)} JSON file(s)")
    
    file_task_ids = {}
    file_task_details = {}
    error_files = []
    no_tasks_files = []
    
    # Extract task information
    print_section("EXTRACTING TASK INFORMATION")
    
    for json_file in json_files:
        rel_path = os.path.relpath(json_file, directory)
        print(f"  Processing: {rel_path}", end='', flush=True)
        
        data, error = load_json(json_file)
        if error:
            error_files.append(f"{rel_path}: {error}")
            print(f" - ERROR")
            continue
        
        task_ids = extract_task_ids_from_release_pipeline(data)
        task_details = extract_task_details_from_release_pipeline(data)
        
        if task_ids is None:
            no_tasks_files.append(rel_path)
            print(f" - No tasks found")
        else:
            file_task_ids[json_file] = task_ids
            file_task_details[json_file] = task_details
            unique_task_ids = set(task_ids)
            print(f" - {len(task_ids)} task(s) ({len(unique_task_ids)} unique)")
    
    print(f"\nSummary:")
    print(f"  Files with tasks: {len(file_task_ids)}")
    print(f"  Files with errors: {len(error_files)}")
    print(f"  Files with no tasks: {len(no_tasks_files)}")
    
    # Analyze patterns
    print_section("ANALYZING PATTERNS")
    pattern_groups, unique_files = analyze_files(file_task_ids, file_task_details)
    
    print(f"\nResults:")
    print(f"  Found {len(pattern_groups)} pattern(s) with matching task sets")
    print(f"  Found {len(unique_files)} unique pipeline(s) with no matches")
    
    return pattern_groups, unique_files, file_task_details

def main():
    """Main function"""
    
    print_header("RELEASE PIPELINE PATTERN IDENTIFIER - MINIMAL OUTPUT")
    print("Only showing: Pipeline Name and Pipeline Path")
    
    # Check for openpyxl
    try:
        from openpyxl import Workbook
    except ImportError:
        print("\nError: openpyxl is required for Excel output.")
        print("Install it using: pip install openpyxl")
        sys.exit(1)
    
    # Parse command line argument
    directory = parse_command_line_arguments()
    
    if directory is None:
        sys.exit(1)
    
    # Validate directory
    if not os.path.exists(directory):
        print(f"\nError: Directory '{directory}' does not exist.")
        sys.exit(1)
    if not os.path.isdir(directory):
        print(f"\nError: '{directory}' is not a directory.")
        sys.exit(1)
    
    # Ensure output directories exist
    logs_dir, output_dir = ensure_directories()
    
    # Generate output filename with timestamp
    timestamp = get_timestamp()
    output_file = os.path.join(output_dir, f"Release_Pipeline_Patterns_{timestamp}.xlsx")
    
    print(f"\nConfiguration:")
    print(f"  Input Directory : {directory}")
    print(f"  Output Excel    : {output_file}")
    print(f"  Logs Directory  : {logs_dir}")
    print(f"  Columns Displayed: Pipeline Name, Pipeline Path")
    
    # Compare files
    pattern_groups, unique_files, file_task_details = compare_json_files(directory)
    
    # Create Excel output
    print_section("CREATING EXCEL REPORT")
    create_excel_output(pattern_groups, unique_files, file_task_details, directory, output_file)
    
    print_header("RESULTS SUMMARY")
    
    if pattern_groups:
        print(f"\n✓ Found {len(pattern_groups)} pattern(s) with similar pipelines")
        for i, (task_tuple, files) in enumerate(pattern_groups.items(), 1):
            print(f"    Pattern {i}: {len(files)} pipelines")
    else:
        print(f"\n✗ No similar patterns found")
    
    if unique_files:
        print(f"\n✓ Found {len(unique_files)} unique pipeline(s) with no matches")
    
    print(f"\n✓ Excel file created at: {output_file}")
    print("\n" + "=" * 70)

if __name__ == "__main__":
    main()